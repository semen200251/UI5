"""Модуль отвечает за выгрузку обменной формы из файла project."""

import datetime
import logging
import os
import pickle
import time

import pandas as pd
import pythoncom
import win32com.client as win32
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

import settings.readOF as config


def get_project(path):
    """Открывает файл проекта и возвращает объект проекта."""

    if not os.path.isabs(path):
        logging.warning('%s: Путь до файла проекта не абсолютный', get_project.__name__)
    logging.info('%s: Пытаемся открыть файл проекта', get_project.__name__)
    try:
        msp = win32.Dispatch("MSProject.Application", pythoncom.CoInitialize())
        _abs_path = os.path.abspath(path)
        print(_abs_path)
        msp.FileOpen(_abs_path)
        project = msp.ActiveProject
    except Exception:
        logging.error('%s: Файл проекта не смог открыться', get_project.__name__)
        raise Exception('Не получилось открыть файл проекта')
    logging.info('%s: Файл проекта успешно открылся', get_project.__name__)
    return project, msp


def _get_data_task(t):
    """"Получает значения task из нужных столбцов.

    На вход поступает объект Task из project. Из него достаются
    требуемые данные, ими заполняется список, который возвращается
    для дальнейшего использования.
    """
    arr = []
    try:
        for i in config.ID_COLUMN.keys():
            try:
                data = getattr(t, i)
            except Exception as e:
                arr.append("Ошибка чтения")
                continue
            if isinstance(data, datetime.datetime):
                data = datetime.datetime.date(data)
            arr.append(data)
    except Exception as e:
        print(e)
        logging.error('%s: Неверный идентификатор столбца project', get_project.__name__)
        raise Exception('Неверный идентификатор столбца project')
    return arr


def fill_dataframe(project):
    """Заполняет DataFrame значениями из project.

    На вход поступают объекты projectа и приложения MS Project.
    Формируется dataframe с данными из требуемых столбцов и возвращается
    для дальнейшего использования.
    """
    logging.info('%s: Создаем DataFrame из столбцов объекта проекта', fill_dataframe.__name__)
    if not project:
        logging.error('%s: Не удалось получить объект проекта', fill_dataframe.__name__)
        raise Exception("Объект проекта пустой")
    if not config.ID_COLUMN:
        logging.error('%s: Ключевые столбцы не заданы', fill_dataframe.__name__)
        raise Exception("Ключевые столбцы не заданы")
    task_collection = project.Tasks
    data = pd.DataFrame(columns=config.ID_COLUMN.values())
    try:
        for t in task_collection:
            data.loc[len(data.index)] = _get_data_task(t)
    except Exception:
        logging.error('%s: Неверно заполнен словарь столбцов и их идентификаторов', fill_dataframe.__name__)
        raise Exception("Ошибка в словаре слобцов и их идентификаторов")
    logging.info('%s: DataFrame из столбцов объекта проекта успешно создан', fill_dataframe.__name__)

    return data


def set_style_excel(column_index, path_to_excel):
    """Применяет стили к строкам excel.

    На вход поступают индексы колонок в Excel и путь до Excel.
    Функция создает объект worksheet из объекта workbook и
    применяет изменения из pickle файла, в котором содержится
    словарь для стилей для каждого конкретного ключевого значения
    (Фаза, феха и т.д.)
    """
    try:
        print(os.getcwd())
        with open(config.PATH_TO_STYLE_FILE, 'rb') as file:
            styles_dict = pickle.load(file)
    except FileNotFoundError:
        logging.error('%s: Неверно задан путь к файлу со стилями', fill_dataframe.__name__)
        raise Exception("Неверный путь до файла со стилями")
    workbook_other = openpyxl.load_workbook(path_to_excel)
    worksheet_other = workbook_other.active
    for row in worksheet_other.iter_rows(min_row=1):
        cell = row[column_index - 1]
        cell_value = cell.value
        if cell_value in styles_dict:
            for cell in row:
                style = styles_dict[cell_value]
                cell.style = style
                column_letter = get_column_letter(cell.column)
                text_length = len(str(cell.value))
                current_width = worksheet_other.column_dimensions[column_letter].width
                if text_length > current_width:
                    worksheet_other.column_dimensions[column_letter].width = text_length

    for column_index, column in enumerate(worksheet_other.columns, start=1):
        for cell in column:
            if isinstance(cell.value, datetime.date):
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    workbook_other.save(path_to_excel)


def main(path_to_project, path_to_folder):
    """Управляющая функция.

    На вход поступает путь до файла project и до файла Excel.
    Выполняется выгрузка обменной формы. В качестве результата
    возвращается абсолютный путь до ОФ.
    """
    path_to_excel = None
    try:
        start = time.time()
        project, msp = get_project(path_to_project)
        data = fill_dataframe(project)
        file_name = os.path.splitext(os.path.basename(path_to_project))[0]
        current_date = datetime.datetime.now().strftime("%d.%m.%Y")
        path_to_excel = path_to_folder + "//" + file_name + "_ОФ_" + current_date + ".xlsx"
        data.to_excel(path_to_excel, sheet_name=f"Обменная форма {datetime.date.today()}", index=False)
        column_index = data.columns.get_loc(config.ID_COLUMN['Text5']) + 1
        set_style_excel(column_index, path_to_excel)
        end = time.time()
        print(end - start)
    except Exception as e:
        print(e)
        return path_to_excel
    msp.Quit()
    return path_to_excel

